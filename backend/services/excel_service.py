import io
import json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
SETTINGS_FILL = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
EXAMPLE_FILL = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def _style_settings_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = SETTINGS_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def generate_template() -> bytes:
    """Generate a styled Excel template with example data."""
    wb = Workbook()

    # --- Sheet 1: SDP Integration ---
    ws = wb.active
    ws.title = "SDP Integration"
    ws.cell(1, 1, "SDP Integration Settings").font = Font(bold=True, size=13, color="1A73E8")

    settings_headers = ["Transport", "Port", "Strategy", "Applications", "Initiate Connection", "Raise Alarm"]
    ws.append([])
    ws.append(settings_headers)
    _style_settings_header(ws, 3, len(settings_headers))
    ws.append(["sctp", "3868", "round-robin", "16777232,16777359,16777302,16777304,16777361", "true", "true"])
    for c in range(1, 7):
        ws.cell(4, c).fill = EXAMPLE_FILL
        ws.cell(4, c).border = THIN_BORDER

    ws.append([])
    ws.append([])
    data_headers = ["App Group", "Realm", "SDP IDs", "Idx1 Peer Host", "Idx1 Connect IPs", "Idx2 Peer Host", "Idx2 Connect IPs"]
    ws.append(data_headers)
    _style_header(ws, 7, len(data_headers))
    examples = [
        ["cha1", "sdp01.realm.com", "sdp01.cs.", "sdp01a.realm.com", "10.1.1.1,10.1.1.2", "sdp01c.realm.com", "10.1.1.5,10.1.1.6"],
        ["cha1", "sdp01.realm.com", "sdp01.cs.", "sdp01b.realm.com", "10.1.1.3,10.1.1.4", "sdp01d.realm.com", "10.1.1.7,10.1.1.8"],
        ["cha1", "sdp02.realm.com", "sdp02.cs.", "sdp02a.realm.com", "10.2.1.1,10.2.1.2", "", ""],
        ["cha2", "sdp01.realm.com", "sdp01.cs.", "sdp01a.realm.com", "10.1.1.1,10.1.1.2", "", ""],
    ]
    for row in examples:
        ws.append(row)
    _auto_width(ws)

    # --- Sheet 2: Subscriber Account Location ---
    ws2 = wb.create_sheet("Subscriber Account Location")
    ws2.cell(1, 1, "Subscriber Account Location").font = Font(bold=True, size=13, color="1A73E8")
    ws2.append([])
    headers2 = ["SDP Name", "IP Address", "Partition ID"]
    ws2.append(headers2)
    _style_header(ws2, 3, len(headers2))
    ws2.append(["sdp01.cs.", "10.0.0.1", "1"])
    ws2.append(["sdp02.cs.", "10.0.0.2", "1"])
    _auto_width(ws2)

    # --- Sheet 3: Diameter ---
    ws3 = wb.create_sheet("Diameter")
    ws3.cell(1, 1, "Diameter Integration").font = Font(bold=True, size=13, color="1A73E8")
    ws3.append([])
    s_headers = ["App Group", "Transport", "Port", "Scheme", "Initiate", "Raise Alarm"]
    ws3.append(s_headers)
    _style_settings_header(ws3, 3, len(s_headers))
    ws3.append(["cha1", "tcp", "3868", "aaa", "true", "true"])
    ws3.append([])
    ws3.append([])
    d_headers = ["App Group", "Interface", "Host", "Realm", "Connect IPs"]
    ws3.append(d_headers)
    _style_header(ws3, 7, len(d_headers))
    ws3.append(["cha1", "Gy", "peer1.com", "realm1.com", "10.0.0.1,10.0.0.2"])
    _auto_width(ws3)

    # --- Sheet 4: NRF Servers ---
    ws4 = wb.create_sheet("NRF Servers")
    ws4.cell(1, 1, "NRF Servers").font = Font(bold=True, size=13, color="1A73E8")
    ws4.append([])
    headers4 = ["Address", "Secured", "Compression", "App Group", "Failure Codes", "NF Service Type"]
    ws4.append(headers4)
    _style_header(ws4, 3, len(headers4))
    ws4.append(["https://nrf.example.com:3002", "true", "true", "global", "404,500", "nfservices"])
    _auto_width(ws4)

    # --- Sheet 5: NRF OAuth Servers ---
    ws5 = wb.create_sheet("NRF OAuth Servers")
    ws5.cell(1, 1, "NRF OAuth Servers").font = Font(bold=True, size=13, color="1A73E8")
    ws5.append([])
    headers5 = ["Address", "Secured", "App Group", "Failure Codes"]
    ws5.append(headers5)
    _style_header(ws5, 3, len(headers5))
    ws5.append(["https://oauth.example.com:3002", "true", "global", "404,500"])
    _auto_width(ws5)

    # --- Sheet 6: NF Profile ---
    ws6 = wb.create_sheet("NF Profile")
    ws6.cell(1, 1, "NF Profile Configuration").font = Font(bold=True, size=13, color="1A73E8")
    ws6.append([])
    headers6 = ["App Group", "NF Type", "NF Status", "HeartBeat Timer", "FQDN", "IPv4 Addresses",
                "Priority", "Capacity", "Allowed NF Types", "PLMN List (JSON)", "S-NSSAIs (JSON)",
                "CHF Info (JSON)", "Custom Info (JSON)"]
    ws6.append(headers6)
    _style_header(ws6, 3, len(headers6))
    ws6.append(["global", "CHF", "REGISTERED", "60", "chf.example.com", "10.0.0.1,10.0.0.2",
                "0", "100", "SMF,SMSF,PCF", '[{"mcc":"240","mnc":"01"}]', '[{"sst":1}]', "", ""])
    _auto_width(ws6)

    # --- Sheet 7: ENM Integration ---
    ws7 = wb.create_sheet("ENM")
    ws7.cell(1, 1, "ENM SNMP Alarm Integration").font = Font(bold=True, size=13, color="1A73E8")
    ws7.append([])
    headers7 = ["OAM Ingress IP", "ENM FM VIP", "ENM Port", "SNMP Version", "Community",
                "Username", "Security Level", "Auth Protocol", "Auth Password", "Priv Protocol", "Priv Password"]
    ws7.append(headers7)
    _style_header(ws7, 3, len(headers7))
    ws7.append(["10.0.0.100", "10.0.0.200", "162", "v2c", "public", "", "", "", "", "", ""])
    _auto_width(ws7)

    # --- Sheet 8: Syslog ---
    ws8 = wb.create_sheet("Syslog")
    ws8.cell(1, 1, "Syslog Egress Configuration").font = Font(bold=True, size=13, color="1A73E8")
    ws8.append([])
    headers8 = ["Host", "Port", "Protocol", "TLS Enabled", "Trust List Name", "Inclusions (JSON)"]
    ws8.append(headers8)
    _style_header(ws8, 3, len(headers8))
    ws8.append(["syslog.example.com", "514", "udp", "false", "", '[{"field":"log_type","value":"audit"}]'])
    _auto_width(ws8)

    # --- Sheet 9: Mediation (EDM Destinations) ---
    ws9 = wb.create_sheet("Mediation")
    ws9.cell(1, 1, "Mediation - Publishing Destinations").font = Font(bold=True, size=13, color="1A73E8")
    ws9.append([])

    s_headers = ["Retry Attempts", "Binary Mode", "Strict Host Key", "Conn Timeout (ms)"]
    ws9.append(s_headers)
    _style_settings_header(ws9, 3, len(s_headers))
    ws9.append(["0", "false", "yes", "5000"])
    for c in range(1, 5):
        ws9.cell(4, c).fill = EXAMPLE_FILL
        ws9.cell(4, c).border = THIN_BORDER

    ws9.append([])
    ws9.append([])
    headers9 = ["Partition ID", "File Type", "Primary Host", "Primary Port", "Primary User",
                "Primary Password", "Primary Dest Folder", "Primary Error Folder",
                "Secondary Host", "Secondary Port", "Secondary User", "Secondary Password",
                "Secondary Dest Folder", "Secondary Error Folder"]
    ws9.append(headers9)
    _style_header(ws9, 7, len(headers9))
    ws9.append(["1", "CDR", "sftp.example.com", "22", "sftpuser", "pass123",
                "/upload/cdr/", "/upload/cdr_error/",
                "sftp-bkp.example.com", "22", "sftpuser", "pass123",
                "/upload/cdr/", "/upload/cdr_error/"])
    ws9.append(["1", "SNAPSHOT", "sftp.example.com", "22", "sftpuser", "pass123",
                "/upload/snapshot/", "/upload/snapshot_error/", "", "", "", "", "", ""])
    _auto_width(ws9)

    # --- Sheet 10: Certificate Mappings ---
    ws10 = wb.create_sheet("Certificates")
    ws10.cell(1, 1, "Certificate Service Mappings").font = Font(bold=True, size=13, color="1A73E8")
    ws10.cell(2, 1, "Note: Actual cert files must be uploaded via UI. This sheet defines mappings only.").font = Font(italic=True, color="90A4AE")
    ws10.append([])
    headers10 = ["Service Name", "Key Name", "Cert Name", "Trust List Name"]
    ws10.append(headers10)
    _style_header(ws10, 4, len(headers10))
    ws10.append(["NRF SBI", "nrf-sbi-key", "nrf-sbi-cert", "nrf-trusted-ca"])
    ws10.append(["SCP SBI", "scp-sbi-key", "scp-sbi-cert", "scp-trusted-ca"])
    _auto_width(ws10)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_current_config(realms, peers, sub_acct_locs, partitions, nrf_servers, oauth_servers, nf_profiles) -> bytes:
    """Export current cluster config to styled Excel."""
    wb = Workbook()

    # --- SDP ---
    ws = wb.active
    ws.title = "SDP Integration"
    ws.cell(1, 1, "SDP Integration - Current Config").font = Font(bold=True, size=13, color="1A73E8")
    ws.append([])
    headers = ["App Group", "Realm", "SDP IDs", "Idx1 Peer Host", "Idx1 Connect IPs", "Idx2 Peer Host", "Idx2 Connect IPs"]
    ws.append(headers)
    _style_header(ws, 3, len(headers))

    def strip_uri(uri):
        return uri.replace("aaa://", "").replace("aaas://", "").split(";")[0]

    def find_peer_connect(uri, app_grp):
        for p in (peers or []):
            if p.get("peer", "").lower() == uri.lower() and p.get("appGrp") == app_grp:
                return ",".join(p.get("connectAddresses", []))
        return ""

    for r in (realms or []):
        idx1 = next((a for a in r.get("addresses", []) if a.get("index") == 1), None)
        idx2 = next((a for a in r.get("addresses", []) if a.get("index") == 2), None)
        idx1_peers = idx1.get("peerAddresses", []) if idx1 else []
        idx2_peers = idx2.get("peerAddresses", []) if idx2 else []
        max_rows = max(len(idx1_peers), len(idx2_peers), 1)
        for i in range(max_rows):
            p1 = idx1_peers[i] if i < len(idx1_peers) else ""
            p2 = idx2_peers[i] if i < len(idx2_peers) else ""
            ws.append([
                r.get("appGrp", ""),
                r.get("realm", "") if i == 0 else "",
                ",".join(r.get("sdp_id", [])) if i == 0 else "",
                strip_uri(p1) if p1 else "",
                find_peer_connect(p1, r.get("appGrp", "")) if p1 else "",
                strip_uri(p2) if p2 else "",
                find_peer_connect(p2, r.get("appGrp", "")) if p2 else "",
            ])
    _auto_width(ws)

    # --- Sub Acct Loc ---
    ws2 = wb.create_sheet("Subscriber Account Location")
    ws2.cell(1, 1, "Subscriber Account Location - Current").font = Font(bold=True, size=13, color="1A73E8")
    ws2.append([])
    ws2.append(["SDP Name", "IP Address", "Partition ID"])
    _style_header(ws2, 3, 3)
    for s in (sub_acct_locs or []):
        ws2.append([s.get("name", ""), s.get("ip", ""), s.get("partitionId", "")])
    _auto_width(ws2)

    # --- NRF Servers ---
    ws4 = wb.create_sheet("NRF Servers")
    ws4.cell(1, 1, "NRF Servers - Current").font = Font(bold=True, size=13, color="1A73E8")
    ws4.append([])
    ws4.append(["Address", "Secured", "Compression", "App Group", "Failure Codes"])
    _style_header(ws4, 3, 5)
    if isinstance(nrf_servers, dict):
        for k, v in nrf_servers.items():
            ws4.append([v.get("address", ""), str(v.get("secured", "")), str(v.get("compression", "")),
                        v.get("appGrp", ""), ",".join(str(c) for c in v.get("failureCodes", []))])
    _auto_width(ws4)

    # --- OAuth Servers ---
    ws5 = wb.create_sheet("NRF OAuth Servers")
    ws5.cell(1, 1, "NRF OAuth Servers - Current").font = Font(bold=True, size=13, color="1A73E8")
    ws5.append([])
    ws5.append(["Address", "Secured", "App Group", "Failure Codes"])
    _style_header(ws5, 3, 4)
    if isinstance(oauth_servers, dict):
        for k, v in oauth_servers.items():
            ws5.append([v.get("address", ""), str(v.get("secured", "")),
                        v.get("appGrp", ""), ",".join(str(c) for c in v.get("failureCodes", []))])
    _auto_width(ws5)

    # --- NF Profile ---
    ws6 = wb.create_sheet("NF Profile")
    ws6.cell(1, 1, "NF Profile - Current").font = Font(bold=True, size=13, color="1A73E8")
    ws6.append([])
    ws6.append(["App Group", "NF Type", "NF Status", "HeartBeat Timer", "FQDN", "IPv4 Addresses", "Priority", "Capacity", "Allowed NF Types"])
    _style_header(ws6, 3, 9)
    if isinstance(nf_profiles, dict):
        for k, v in nf_profiles.items():
            ws6.append([k, v.get("nfType", ""), v.get("nfStatus", ""), v.get("heartBeatTimer", ""),
                        v.get("fqdn", ""), ",".join(v.get("ipv4Addresses", [])),
                        v.get("priority", ""), v.get("capacity", ""), ",".join(v.get("allowedNfTypes", []))])
    _auto_width(ws6)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_excel(file_bytes: bytes) -> dict:
    """Parse uploaded Excel and return structured data for all integrations."""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    result = {}

    # --- SDP Integration ---
    if "SDP Integration" in wb.sheetnames:
        ws = wb["SDP Integration"]
        rows = list(ws.iter_rows(values_only=True))
        settings = {}
        data_rows = []
        settings_found = False
        data_found = False

        for i, row in enumerate(rows):
            if row and row[0] and str(row[0]).strip().lower() == "transport":
                settings_found = True
                continue
            if settings_found and not settings and row and row[0]:
                settings = {
                    "transport": str(row[0] or "sctp"),
                    "port": str(row[1] or "3868"),
                    "strategy": str(row[2] or "round-robin"),
                    "applications": str(row[3] or ""),
                    "initiateConnection": str(row[4] or "true").lower() == "true",
                    "raiseAlarm": str(row[5] or "true").lower() == "true",
                }
                settings_found = False
                continue
            if row and row[0] and str(row[0]).strip().lower() == "app group":
                data_found = True
                continue
            if data_found and row and row[0]:
                data_rows.append({
                    "appGrp": str(row[0] or ""),
                    "realm": str(row[1] or ""),
                    "sdpIds": str(row[2] or ""),
                    "idx1Host": str(row[3] or ""),
                    "idx1Connect": str(row[4] or ""),
                    "idx2Host": str(row[5] or "") if len(row) > 5 else "",
                    "idx2Connect": str(row[6] or "") if len(row) > 6 else "",
                })

        result["sdp"] = {"settings": settings, "entries": data_rows}

    # --- Subscriber Account Location ---
    if "Subscriber Account Location" in wb.sheetnames:
        ws = wb["Subscriber Account Location"]
        rows = list(ws.iter_rows(values_only=True))
        entries = []
        data_found = False
        for row in rows:
            if row and row[0] and str(row[0]).strip().lower() == "sdp name":
                data_found = True
                continue
            if data_found and row and row[0]:
                entries.append({"name": str(row[0]), "ip": str(row[1] or ""), "partitionId": str(row[2] or "1")})
        result["subAcctLoc"] = entries

    # --- Diameter ---
    if "Diameter" in wb.sheetnames:
        ws = wb["Diameter"]
        rows = list(ws.iter_rows(values_only=True))
        settings = {}
        entries = []
        settings_found = False
        data_found = False
        for row in rows:
            if row and row[0] and str(row[0]).strip().lower() == "app group" and not data_found:
                if not settings:
                    settings_found = True
                else:
                    data_found = True
                continue
            if settings_found and row and row[0]:
                settings = {
                    "appGrp": str(row[0] or ""),
                    "transport": str(row[1] or "tcp"),
                    "port": str(row[2] or "3868"),
                    "scheme": str(row[3] or "aaa"),
                    "initiate": str(row[4] or "true").lower() == "true",
                    "raiseAlarm": str(row[5] or "true").lower() == "true",
                }
                settings_found = False
                continue
            if data_found and row and row[0]:
                entries.append({
                    "appGrp": str(row[0] or ""),
                    "interface": str(row[1] or ""),
                    "host": str(row[2] or ""),
                    "realm": str(row[3] or ""),
                    "connectIps": str(row[4] or "") if len(row) > 4 else "",
                })
        result["diameter"] = {"settings": settings, "entries": entries}

    # --- NRF Servers ---
    if "NRF Servers" in wb.sheetnames:
        ws = wb["NRF Servers"]
        rows = list(ws.iter_rows(values_only=True))
        entries = []
        data_found = False
        for row in rows:
            if row and row[0] and str(row[0]).strip().lower() == "address":
                data_found = True
                continue
            if data_found and row and row[0]:
                entries.append({
                    "address": str(row[0]),
                    "secured": str(row[1] or "true").lower() == "true",
                    "compression": str(row[2] or "true").lower() == "true" if len(row) > 2 else True,
                    "appGrp": str(row[3] or "global") if len(row) > 3 else "global",
                    "failureCodes": [int(x.strip()) for x in str(row[4] or "").split(",") if x.strip().isdigit()] if len(row) > 4 else [],
                    "nfServiceType": str(row[5] or "nfservices") if len(row) > 5 else "nfservices",
                })
        result["nrfServers"] = entries

    # --- NRF OAuth Servers ---
    if "NRF OAuth Servers" in wb.sheetnames:
        ws = wb["NRF OAuth Servers"]
        rows = list(ws.iter_rows(values_only=True))
        entries = []
        data_found = False
        for row in rows:
            if row and row[0] and str(row[0]).strip().lower() == "address":
                data_found = True
                continue
            if data_found and row and row[0]:
                entries.append({
                    "address": str(row[0]),
                    "secured": str(row[1] or "true").lower() == "true",
                    "appGrp": str(row[2] or "global") if len(row) > 2 else "global",
                    "failureCodes": [int(x.strip()) for x in str(row[3] or "").split(",") if x.strip().isdigit()] if len(row) > 3 else [],
                })
        result["oauthServers"] = entries

    # --- NF Profile ---
    if "NF Profile" in wb.sheetnames:
        ws = wb["NF Profile"]
        rows = list(ws.iter_rows(values_only=True))
        entries = []
        data_found = False
        for row in rows:
            if row and row[0] and str(row[0]).strip().lower() == "app group":
                data_found = True
                continue
            if data_found and row and row[0]:
                profile = {"appGroup": str(row[0]), "nfType": str(row[1] or "CHF"), "nfStatus": str(row[2] or "REGISTERED")}
                if len(row) > 3 and row[3]: profile["heartBeatTimer"] = int(row[3])
                if len(row) > 4 and row[4]: profile["fqdn"] = str(row[4])
                if len(row) > 5 and row[5]: profile["ipv4Addresses"] = [s.strip() for s in str(row[5]).split(",") if s.strip()]
                if len(row) > 6 and row[6]: profile["priority"] = int(row[6])
                if len(row) > 7 and row[7]: profile["capacity"] = int(row[7])
                if len(row) > 8 and row[8]: profile["allowedNfTypes"] = [s.strip() for s in str(row[8]).split(",") if s.strip()]
                if len(row) > 9 and row[9]:
                    try: profile["plmnList"] = json.loads(str(row[9]))
                    except: pass
                if len(row) > 10 and row[10]:
                    try: profile["sNssais"] = json.loads(str(row[10]))
                    except: pass
                if len(row) > 11 and row[11]:
                    try: profile["chfInfo"] = json.loads(str(row[11]))
                    except: pass
                if len(row) > 12 and row[12]:
                    try: profile["customInfo"] = json.loads(str(row[12]))
                    except: pass
                entries.append(profile)
        result["nfProfile"] = entries

    # --- ENM ---
    if "ENM" in wb.sheetnames:
        ws = wb["ENM"]
        rows = list(ws.iter_rows(values_only=True))
        entries = []
        data_found = False
        for row in rows:
            if row and row[0] and str(row[0]).strip().lower() == "oam ingress ip":
                data_found = True
                continue
            if data_found and row and row[0]:
                entry = {
                    "oamIngressIp": str(row[0]),
                    "enmFmVip": str(row[1] or ""),
                    "enmPort": str(row[2] or "162"),
                    "version": str(row[3] or "v2c"),
                    "community": str(row[4] or "public") if len(row) > 4 else "public",
                }
                if len(row) > 5 and row[5]: entry["userName"] = str(row[5])
                if len(row) > 6 and row[6]: entry["securityLevel"] = str(row[6])
                if len(row) > 7 and row[7]: entry["authProtocol"] = str(row[7])
                if len(row) > 8 and row[8]: entry["authPassword"] = str(row[8])
                if len(row) > 9 and row[9]: entry["privProtocol"] = str(row[9])
                if len(row) > 10 and row[10]: entry["privPassword"] = str(row[10])
                entries.append(entry)
        result["enm"] = entries

    # --- Syslog ---
    if "Syslog" in wb.sheetnames:
        ws = wb["Syslog"]
        rows = list(ws.iter_rows(values_only=True))
        entries = []
        data_found = False
        for row in rows:
            if row and row[0] and str(row[0]).strip().lower() == "host":
                data_found = True
                continue
            if data_found and row and row[0]:
                entry = {
                    "host": str(row[0]),
                    "port": str(row[1] or "514"),
                    "protocol": str(row[2] or "udp"),
                    "tlsEnabled": str(row[3] or "false").lower() == "true" if len(row) > 3 else False,
                    "trustListName": str(row[4] or "") if len(row) > 4 else "",
                }
                if len(row) > 5 and row[5]:
                    try: entry["inclusions"] = json.loads(str(row[5]))
                    except: pass
                entries.append(entry)
        result["syslog"] = entries

    # --- Mediation ---
    if "Mediation" in wb.sheetnames:
        ws = wb["Mediation"]
        rows = list(ws.iter_rows(values_only=True))
        settings = {}
        entries = []
        settings_found = False
        data_found = False
        for row in rows:
            if row and row[0] and str(row[0]).strip().lower() == "retry attempts":
                settings_found = True
                continue
            if settings_found and not settings and row and row[0] is not None:
                settings = {
                    "retryAttempts": str(row[0] or "0"),
                    "binaryMode": str(row[1] or "false"),
                    "strictHostKey": str(row[2] or "yes"),
                    "connTimeout": str(row[3] or "5000"),
                }
                settings_found = False
                continue
            if row and row[0] and str(row[0]).strip().lower() == "partition id":
                data_found = True
                continue
            if data_found and row and row[0] is not None:
                entry = {
                    "partitionId": str(row[0]),
                    "fileType": str(row[1] or "") if len(row) > 1 else "",
                    "primaryHost": str(row[2] or "") if len(row) > 2 else "",
                    "primaryPort": str(row[3] or "22") if len(row) > 3 else "22",
                    "primaryUser": str(row[4] or "") if len(row) > 4 else "",
                    "primaryPassword": str(row[5] or "") if len(row) > 5 else "",
                    "primaryDestFolder": str(row[6] or "") if len(row) > 6 else "",
                    "primaryErrorFolder": str(row[7] or "") if len(row) > 7 else "",
                    "secondaryHost": str(row[8] or "") if len(row) > 8 else "",
                    "secondaryPort": str(row[9] or "22") if len(row) > 9 else "22",
                    "secondaryUser": str(row[10] or "") if len(row) > 10 else "",
                    "secondaryPassword": str(row[11] or "") if len(row) > 11 else "",
                    "secondaryDestFolder": str(row[12] or "") if len(row) > 12 else "",
                    "secondaryErrorFolder": str(row[13] or "") if len(row) > 13 else "",
                }
                entries.append(entry)
        result["mediation"] = {"settings": settings, "entries": entries}

    # --- Certificates ---
    if "Certificates" in wb.sheetnames:
        ws = wb["Certificates"]
        rows = list(ws.iter_rows(values_only=True))
        entries = []
        data_found = False
        for row in rows:
            if row and row[0] and str(row[0]).strip().lower() == "service name":
                data_found = True
                continue
            if data_found and row and row[0]:
                entries.append({
                    "serviceName": str(row[0]),
                    "keyName": str(row[1] or ""),
                    "certName": str(row[2] or ""),
                    "trustListName": str(row[3] or ""),
                })
        result["certificates"] = entries

    return result
